from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import re
from rapidfuzz import process, fuzz
import psycopg2
from psycopg2.extras import execute_values

# URLs ตลาด (export เป็น CSV)
TALATTHAI_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRSZiUoIiB8i-QmXDTC2AWASSYr7gMaLjCx4HSikqeKN_9-ZSFpsHb6GVRLwR7o9gnkCHkulx5Gii2z"
    "/pub?output=csv&gid=1914406586"
)
# เปลี่ยนตรงนี้ให้เป็น CSV URL ใหม่
YINGCHAROEN_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRmFKGmbryvyQ2SJ6dfgtr3U9QfuECaOIKkB_dly-Jkm6j7hu92tpHbXZyQtHeSHgOWaBT9FIU6jsga"
    "/pub?output=csv&gid=1711908758"
)


def load_csv_df(url: str) -> pd.DataFrame:
    return pd.read_csv(url)

def get_db_connection():
    return psycopg2.connect(
        host="localhost",
        database="postgres",
        user="postgres",
        password="Earth32471"
    )

def safe_num(x):
    s = str(x).strip().replace(",", "")
    if s in ("", "-", "nan", "NaN", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None

def insert_data_to_db(df: pd.DataFrame, item_col: str):
    # แปลงคอลัมน์ตัวเลข
    for col in ["ปริมาณเฉลี่ย", "ราคา", "ราคา_ตลาดไท", "ราคา_ยิ่งเจริญ"]:
        if col in df.columns:
            df[col] = df[col].apply(safe_num)

    # ตัดแถวซ้ำกันตาม sku (item_col) ให้เหลือแถวแรกของแต่ละ sku
    df = df.drop_duplicates(subset=[item_col], keep="first")

    conn = get_db_connection()
    cur = conn.cursor()
    insert_query = """
    INSERT INTO public.market_data (
      sku, หน่วย, ปริมาณเฉลี่ย, ราคา,
      ชื่อสินค้า_clean, ชื่อตลาดไท, ราคา_ตลาดไท, หน่วย_ตลาดไท, รูป_ตลาดไท,
      ชื่อยิ่งเจริญ, ราคา_ยิ่งเจริญ, หน่วย_ยิ่งเจริญ, รูป_ยิ่งเจริญ
    ) VALUES %s
    ON CONFLICT (sku) DO UPDATE
      SET
        หน่วย             = EXCLUDED.หน่วย,
        ปริมาณเฉลี่ย       = EXCLUDED.ปริมาณเฉลี่ย,
        ราคา              = EXCLUDED.ราคา,
        ชื่อสินค้า_clean   = EXCLUDED.ชื่อสินค้า_clean,
        ชื่อตลาดไท        = EXCLUDED.ชื่อตลาดไท,
        ราคา_ตลาดไท       = EXCLUDED.ราคา_ตลาดไท,
        หน่วย_ตลาดไท      = EXCLUDED.หน่วย_ตลาดไท,
        รูป_ตลาดไท        = EXCLUDED.รูป_ตลาดไท,
        ชื่อยิ่งเจริญ      = EXCLUDED.ชื่อยิ่งเจริญ,
        ราคา_ยิ่งเจริญ     = EXCLUDED.ราคา_ยิ่งเจริญ,
        หน่วย_ยิ่งเจริญ    = EXCLUDED.หน่วย_ยิ่งเจริญ,
        รูป_ยิ่งเจริญ      = EXCLUDED.รูป_ยิ่งเจริญ;
    """

    values = [
        (
            row[item_col],               # sku
            row.get("หน่วย"),
            row.get("ปริมาณเฉลี่ย"),
            row.get("ราคา"),
            row.get("ชื่อสินค้า_clean"),
            row.get("ชื่อตลาดไท"),
            row.get("ราคา_ตลาดไท"),
            row.get("หน่วย_ตลาดไท"),
            row.get("รูป_ตลาดไท"),
            row.get("ชื่อยิ่งเจริญ"),
            row.get("ราคา_ยิ่งเจริญ"),
            row.get("หน่วย_ยิ่งเจริญ"),
            row.get("รูป_ยิ่งเจริญ"),
        )
        for _, row in df.iterrows()
    ]

    if values:
        execute_values(cur, insert_query, values)
        conn.commit()

    cur.close()
    conn.close()

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# โหลด DataFrame ตลาดล่วงหน้า
df_talatthai   = load_csv_df(TALATTHAI_CSV_URL)
df_yingcharoen = load_csv_df(YINGCHAROEN_CSV_URL)

def extract_thai(text: str) -> str:
    return "".join(re.findall(r"[\u0E00-\u0E7F]+", str(text)))

def match_price_unit(df_market: pd.DataFrame, name: str, threshold=80):
    best_match, best_score = None, 0
    for m in df_market.get("ชื่อสินค้า", []):
        score = fuzz.partial_ratio(name, m)
        if score >= threshold and score > best_score:
            best_score, best_match = score, m
    if not best_match:
        return None, None, None, None

    row = df_market[df_market["ชื่อสินค้า"] == best_match].iloc[0]
    raw_price = str(row.get("ราคา", "")).replace(",", "")
    m = re.search(r"([\d\.]+)", raw_price)
    price_numeric = float(m.group(1)) if m else None
    unit = row.get("หน่วย")
    url = row.get("URL รูป", "")
    link = f'=HYPERLINK("{url}","เปิดรูป")' if url else ""
    return row["ชื่อสินค้า"], price_numeric, unit, link

def find_header_row(df: pd.DataFrame, possible_cols=["item","รายการ","ชื่อสินค้า"]):
    for i, row in df.iterrows():
        cells = [str(c).strip().lower() for c in row.values]
        if any(col.lower() in cells for col in possible_cols):
            return i
    return None

@app.post("/upload")
async def upload_demand(file: UploadFile = File(...)):
    if not file.filename.lower().endswith((".xls", ".xlsx")):
        raise HTTPException(400, "โปรดอัปโหลดไฟล์ Excel (.xls/.xlsx)")

    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), read_only=True)
    visible = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]
    wb.close()
    if not visible:
        raise HTTPException(400, "ไฟล์ไม่มีชีทที่มองเห็นได้")

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wrote = False
        for sheet in visible:
            df_raw = pd.read_excel(BytesIO(content), sheet_name=sheet, header=None)
            hr = find_header_row(df_raw)
            if hr is None:
                continue

            headers = df_raw.iloc[hr].tolist()
            match = (
                process.extractOne("Item", headers, score_cutoff=90)
                or process.extractOne("รายการ", headers, score_cutoff=90)
                or process.extractOne("ชื่อสินค้า", headers, score_cutoff=90)
            )
            if not match:
                continue

            item_col = match[0]
            df = pd.DataFrame(df_raw.values[hr+1:], columns=headers)
            df["ชื่อสินค้า_clean"] = df[item_col].apply(extract_thai)
            df[["ชื่อตลาดไท","ราคา_ตลาดไท","หน่วย_ตลาดไท","รูป_ตลาดไท"]] = (
                df["ชื่อสินค้า_clean"].apply(lambda x: pd.Series(match_price_unit(df_talatthai, x)))
            )
            df[["ชื่อยิ่งเจริญ","ราคา_ยิ่งเจริญ","หน่วย_ยิ่งเจริญ","รูป_ยิ่งเจริญ"]] = (
                df["ชื่อสินค้า_clean"].apply(lambda x: pd.Series(match_price_unit(df_yingcharoen, x)))
            )

            insert_data_to_db(df, item_col)
            df.to_excel(writer, sheet_name=sheet[:31], index=False)
            wrote = True

        if not wrote:
            pd.DataFrame().to_excel(writer, sheet_name="Sheet1", index=False)

        for ws in writer.book.worksheets:
            ws.sheet_state = "visible"
        writer.book.active = 0

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=result.xlsx"},
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
